#!/usr/bin/env python3
"""Validate workbook structure and key formulas for Link Estimate System."""

from __future__ import annotations

import argparse
import re
import zipfile
from pathlib import Path

import openpyxl


REQUIRED_SHEETS = [
    "操作パネル",
    "ダッシュボード",
    "案件管理",
    "Ｓ表紙",
    "P-003_吉野様邸キッチン",
    "発注書兼請求書",
    "領収書",
    "請求管理",
    "支払管理",
    "工程表",
    "顧客マスタ",
    "マスター設定",
    "工事項目DB",
    "業者DB",
    "案件テンプレート",
    "設計書",
    "請求書",
    "請書 (2)",
    "実行予算　雛形",
    "注文書(控)",
    "注文書",
    "請書",
    "解体工事",
    "産廃費",
    "仮設",
    "解体後補修箇所",
    "新規壁工事",
    "造作工事",
    "下地工事",
    "建具工事",
    "家具",
    "床",
    "電気設備‘",
    "空調",
    "タイル左官工事",
    "雑工事",
    "シート工事",
    "仕上げ工事",
    "衛生設備",
    "塗装工事",
    "防災工事",
    "ガス工事",
    "Claude Log",
    "設計図v2",
    "設計図v3",
]

EXPECTED_HEADERS = {
    1: "案件ID",
    2: "顧客ID",
    3: "顧客名",
    4: "案件名",
    5: "工事内容",
    6: "見積金額",
    7: "実行予算",
    8: "粗利率",
    9: "ステータス",
    10: "ステータス変更日",
    11: "経過日数",
    12: "滞留アラート",
    13: "担当者",
    14: "受注日",
    15: "着工予定日",
    16: "完工日",
    17: "請求予定額",
    18: "失注理由",
    19: "作成日",
    20: "備考",
    21: "完工予定日",
    22: "工期余日",
    23: "工期アラート",
    24: "請求書ステータス",
    25: "請求日",
    26: "請求金額",
    27: "入金ステータス",
    28: "入金日",
    29: "入金額",
    30: "受注区分",
    31: "施工住所",
    32: "物件名",
}


def normalize_formula(value: str) -> str:
    return value.replace(" ", "").upper()


def validate(path: Path, require_vba: bool) -> int:
    errors: list[str] = []
    warnings: list[str] = []

    if not path.exists():
        print(f"[ERROR] Workbook not found: {path}")
        return 1

    wb = openpyxl.load_workbook(path, data_only=False)

    missing = [name for name in REQUIRED_SHEETS if name not in wb.sheetnames]
    if missing:
        errors.append(f"Missing sheets: {', '.join(missing)}")

    if len(wb.sheetnames) != 45:
        errors.append(f"Unexpected sheet count: {len(wb.sheetnames)} (expected 45)")

    s_cover = wb["Ｓ表紙"]
    for offset, row in enumerate(range(36, 56)):
        expected = f"=INDIRECT(\"'\"&$J$2&\"'!B{105 + offset}\")"
        actual = s_cover.cell(row=row, column=9).value or ""
        if normalize_formula(str(actual)) != normalize_formula(expected):
            errors.append(
                f"S表紙!I{row} formula mismatch: actual={actual!r}, expected={expected!r}"
            )

    projects = wb["案件管理"]
    for col, expected_header in EXPECTED_HEADERS.items():
        actual_header = projects.cell(row=4, column=col).value
        if actual_header != expected_header:
            errors.append(
                f"案件管理!{openpyxl.utils.get_column_letter(col)}4 header mismatch: "
                f"actual={actual_header!r}, expected={expected_header!r}"
            )

    indexed_formula_hits: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=") and re.search(r"\[\d+\]", value):
                    indexed_formula_hits.append(f"{ws.title}!{cell.coordinate}")

    if indexed_formula_hits:
        errors.append(
            "Indexed external-like formulas remain: "
            + ", ".join(indexed_formula_hits[:15])
        )

    with zipfile.ZipFile(path) as zf:
        has_vba = any(name.lower().endswith("vbaproject.bin") for name in zf.namelist())
        is_macro_enabled = "macroEnabled" in zf.read("[Content_Types].xml").decode("utf-8")
        external_link_files = [
            name for name in zf.namelist() if name.startswith("xl/externalLinks/")
        ]

    if not is_macro_enabled:
        errors.append("Workbook content types are not macro-enabled.")
    if external_link_files:
        errors.append(
            "Workbook still contains externalLink parts: "
            + ", ".join(external_link_files[:10])
        )

    if require_vba and not has_vba:
        errors.append("Workbook does not contain vbaProject.bin.")
    if not require_vba and not has_vba:
        warnings.append("Workbook does not contain vbaProject.bin (run manual VBA import).")

    for msg in warnings:
        print(f"[WARN] {msg}")
    for msg in errors:
        print(f"[ERROR] {msg}")

    if errors:
        print(f"[FAIL] Validation failed for: {path}")
        return 1

    print(f"[PASS] Validation passed for: {path}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate Link Estimate workbook.")
    parser.add_argument(
        "--workbook",
        default="excel/見積原価管理システム.xlsm",
        help="Path to workbook (.xlsm or .xlsx)",
    )
    parser.add_argument(
        "--require-vba",
        action="store_true",
        help="Fail validation when vbaProject.bin is missing.",
    )
    args = parser.parse_args()

    return validate(Path(args.workbook).resolve(), require_vba=args.require_vba)


if __name__ == "__main__":
    raise SystemExit(main())
