#!/usr/bin/env python3
"""Build macro-enabled workbook artifact from the source .xlsx file.

This script does two things in Excel for Mac:
1. Updates `Ｓ表紙!I36:I55` to dynamic INDIRECT formulas.
2. Saves the workbook as `.xlsm` (macro-enabled file format).

Notes:
- VBA module import and button assignment remain manual Excel UI steps.
- Requires macOS Excel + `appscript` Python package.
"""

from __future__ import annotations

import argparse
import importlib.util
import re
import shutil
import tempfile
from pathlib import Path


def _load_excel_terms():
    from appscript import terminology

    temp_dir = tempfile.mkdtemp(prefix="excelglue-")
    glue_path = Path(temp_dir) / "excel_glue.py"
    terminology.dump("/Applications/Microsoft Excel.app", str(glue_path), usesdef=True)

    spec = importlib.util.spec_from_file_location("excel_glue_runtime", glue_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("Failed to load generated Excel terminology module.")

    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _collect_indexed_formula_fixes(source_xlsx: Path) -> list[tuple[str, str, str]]:
    import openpyxl

    wb = openpyxl.load_workbook(source_xlsx, data_only=False, keep_links=True)
    fixes: list[tuple[str, str, str]] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=") and re.search(r"\[\d+\]", value):
                    new_value = re.sub(r"\[\d+\]", "", value)
                    if new_value != value:
                        fixes.append((ws.title, cell.coordinate, new_value))

    return fixes


def _break_external_links(workbook, excel, k) -> None:
    for _ in range(10):
        try:
            links = workbook.link_sources(type=k.link_type_Excel_links)
        except Exception:
            links = excel.link_sources(workbook, type=k.link_type_Excel_links)

        if not links or links == k.missing_value:
            return

        if not isinstance(links, (list, tuple)):
            links = [links]

        for link in links:
            try:
                workbook.break_link(name=link, type=k.link_type_Excel_links)
            except Exception:
                excel.break_link(workbook, name=link, type=k.link_type_Excel_links)


def build_workbook(source_xlsx: Path, output_xlsm: Path) -> None:
    if not source_xlsx.exists():
        raise FileNotFoundError(f"Source workbook not found: {source_xlsx}")

    excel_terms = _load_excel_terms()
    formula_fixes = _collect_indexed_formula_fixes(source_xlsx)

    from appscript import app, k
    from mactypes import File

    excel = app("/Applications/Microsoft Excel.app", terms=excel_terms)
    excel.ask_to_update_links.set(False)
    excel.display_alerts.set(False)
    excel.screen_updating.set(False)

    workbook = None
    temp_output = Path(tempfile.mkdtemp(prefix="xlsm-build-")) / output_xlsm.name
    try:
        workbook = excel.open_workbook(workbook_file_name=File(str(source_xlsx)))
        s_cover = workbook.worksheets["Ｓ表紙"]

        for row in range(36, 56):
            target_row = 105 + (row - 36)  # B105 ... B124
            formula = f"=INDIRECT(\"'\"&$J$2&\"'!B{target_row}\")"
            s_cover.cells[f"I{row}"].formula.set(formula)

        # Replace legacy indexed formulas like [1]Sheet!A1 -> Sheet!A1
        for sheet_name, cell_address, formula in formula_fixes:
            workbook.worksheets[sheet_name].cells[cell_address].formula.set(formula)

        # Remove any remaining Excel external link sources.
        _break_external_links(workbook, excel, k)

        if temp_output.exists():
            temp_output.unlink()

        workbook.save_workbook_as(
            filename=File(str(temp_output)),
            file_format=k.macro_enabled_XML_file_format,
        )

        # SaveAs switches active workbook reference to the new file.
        excel.active_workbook.close(saving=k.no)

        output_xlsm.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(temp_output, output_xlsm)
    finally:
        # Best effort cleanup (Excel may already be closed).
        try:
            if workbook is not None:
                workbook.close(saving=k.no)
        except Exception:
            pass
        try:
            excel.quit()
        except Exception:
            pass
        try:
            if temp_output.parent.exists():
                shutil.rmtree(temp_output.parent, ignore_errors=True)
        except Exception:
            pass


def main() -> int:
    parser = argparse.ArgumentParser(description="Build .xlsm from source .xlsx workbook.")
    parser.add_argument(
        "--source",
        default="excel/見積原価管理システム.xlsx",
        help="Path to source .xlsx workbook",
    )
    parser.add_argument(
        "--output",
        default="excel/見積原価管理システム.xlsm",
        help="Path to output .xlsm workbook",
    )
    args = parser.parse_args()

    source = Path(args.source).resolve()
    output = Path(args.output).resolve()

    output.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(source, output)

    print(f"Built workbook: {output}")
    print("Reminder: import VBA module and assign buttons in Excel UI.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
