"""Basic API tests for MVP endpoints."""

from __future__ import annotations

import os
import tempfile
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

TMP_DIR = Path(tempfile.mkdtemp(prefix="link-estimate-api-test-"))
os.environ["APP_DATABASE_URL"] = f"sqlite:///{(TMP_DIR / 'test.db').as_posix()}"

from app.main import app  # noqa: E402


def _create_sync_workbook(path: Path) -> None:
    wb = Workbook()

    ws_customer = wb.active
    ws_customer.title = "顧客マスタ"
    ws_customer.cell(4, 1, "顧客ID")
    ws_customer.cell(4, 3, "会社名 / 氏名")
    ws_customer.cell(5, 1, "C-101")
    ws_customer.cell(5, 3, "同期テスト顧客")
    ws_customer.cell(5, 5, "同期担当")
    ws_customer.cell(5, 18, "アクティブ")

    ws_project = wb.create_sheet("案件管理")
    ws_project.cell(4, 1, "案件ID")
    ws_project.cell(5, 1, "P-101")
    ws_project.cell(5, 2, "C-101")
    ws_project.cell(5, 3, "同期テスト顧客")
    ws_project.cell(5, 4, "同期案件")
    ws_project.cell(5, 8, 0.28)
    ws_project.cell(5, 9, "③見積作成中")
    ws_project.cell(5, 13, "吉野博")
    ws_project.cell(5, 19, "2026-02-17")
    ws_project.cell(5, 31, "東京都江戸川区")

    ws_invoice = wb.create_sheet("請求管理")
    ws_invoice.cell(4, 1, "請求ID")
    ws_invoice.cell(5, 1, "INV-101")
    ws_invoice.cell(5, 2, "P-101")
    ws_invoice.cell(5, 3, "同期案件")
    ws_invoice.cell(5, 4, "同期テスト顧客")
    ws_invoice.cell(5, 7, 500000)
    ws_invoice.cell(5, 12, "sync test")

    ws_item = wb.create_sheet("工事項目DB")
    ws_item.cell(4, 1, "ID")
    ws_item.cell(5, 1, 501)
    ws_item.cell(5, 2, "解体工事")
    ws_item.cell(5, 3, "同期明細")
    ws_item.cell(5, 4, "仕様X")
    ws_item.cell(5, 5, "式")
    ws_item.cell(5, 6, 12345)
    ws_item.cell(5, 7, "同期業者")
    ws_item.cell(5, 9, 0.2)

    wb.save(path)


def test_health() -> None:
    with TestClient(app) as client:
        resp = client.get("/health")
        assert resp.status_code == 200
        assert resp.json()["status"] == "ok"


def test_customers() -> None:
    with TestClient(app) as client:
        resp = client.get("/api/v1/customers")
        assert resp.status_code == 200
        data = resp.json()
        assert isinstance(data, list)
        assert len(data) >= 1


def test_create_project_and_list() -> None:
    with TestClient(app) as client:
        create_resp = client.post(
            "/api/v1/projects",
            json={
                "customer_id": "C-001",
                "project_name": "APIテスト案件",
                "site_address": "東京都",
            },
        )
        assert create_resp.status_code == 201
        created = create_resp.json()
        assert created["project_id"].startswith("P-")

        list_resp = client.get("/api/v1/projects")
        assert list_resp.status_code == 200
        payload = list_resp.json()
        assert payload["total"] >= 1


def test_document_exports() -> None:
    with TestClient(app) as client:
        estimate = client.post(
            "/api/v1/documents/estimate-cover",
            json={"project_id": "P-003"},
        )
        assert estimate.status_code == 200
        assert estimate.headers["content-type"] == "application/pdf"

        receipt = client.post(
            "/api/v1/documents/receipt",
            json={"invoice_id": "INV-001"},
        )
        assert receipt.status_code == 200
        assert receipt.headers["content-type"] == "application/pdf"


def test_project_item_endpoints() -> None:
    with TestClient(app) as client:
        master_resp = client.get("/api/v1/work-items")
        assert master_resp.status_code == 200
        masters = master_resp.json()
        assert len(masters) >= 1
        first_master_id = masters[0]["id"]

        create_resp = client.post(
            "/api/v1/projects/P-003/items",
            json={"master_item_id": first_master_id, "quantity": 3},
        )
        assert create_resp.status_code == 201
        created = create_resp.json()
        assert created["line_total"] >= 0

        list_resp = client.get("/api/v1/projects/P-003/items")
        assert list_resp.status_code == 200
        items = list_resp.json()
        assert any(item["id"] == created["id"] for item in items)


def test_excel_sync_endpoint() -> None:
    wb_path = TMP_DIR / "sync_source.xlsx"
    _create_sync_workbook(wb_path)

    with TestClient(app) as client:
        sync_resp = client.post(
            "/api/v1/sync/excel",
            json={"workbook_path": str(wb_path)},
        )
        assert sync_resp.status_code == 200
        body = sync_resp.json()
        assert body["customers_upserted"] >= 1
        assert body["projects_upserted"] >= 1
        assert body["invoices_upserted"] >= 1
        assert body["payments_upserted"] >= 0
        assert body["work_items_upserted"] >= 1

        projects_resp = client.get("/api/v1/projects", params={"customer_id": "C-101"})
        assert projects_resp.status_code == 200
        projects_body = projects_resp.json()
        assert projects_body["total"] >= 1
        assert any(p["project_id"] == "P-101" for p in projects_body["items"])


def test_excel_sync_upload_endpoint() -> None:
    wb_path = TMP_DIR / "sync_upload_source.xlsx"
    _create_sync_workbook(wb_path)

    with TestClient(app) as client:
        with wb_path.open("rb") as fp:
            sync_resp = client.post(
                "/api/v1/sync/excel/upload",
                files={"file": ("sync_upload_source.xlsx", fp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert sync_resp.status_code == 200
        body = sync_resp.json()
        assert body["customers_upserted"] >= 1
        assert body["projects_upserted"] >= 1
        assert body["invoices_upserted"] >= 1
        assert body["work_items_upserted"] >= 1


def test_invoice_and_payment_endpoints() -> None:
    with TestClient(app) as client:
        invoice_create = client.post(
            "/api/v1/invoices",
            json={
                "project_id": "P-003",
                "invoice_amount": 300000,
                "invoice_type": "一括",
                "paid_amount": 100000,
                "status": "⚠一部入金",
            },
        )
        assert invoice_create.status_code == 201
        invoice_body = invoice_create.json()
        assert invoice_body["invoice_id"].startswith("INV-")
        assert invoice_body["remaining_amount"] == 200000

        invoice_list = client.get("/api/v1/invoices", params={"project_id": "P-003"})
        assert invoice_list.status_code == 200
        invoices = invoice_list.json()
        assert any(x["invoice_id"] == invoice_body["invoice_id"] for x in invoices)

        invoice_patch = client.patch(
            f"/api/v1/invoices/{invoice_body['invoice_id']}",
            json={"paid_amount": 300000},
        )
        assert invoice_patch.status_code == 200
        patched_invoice = invoice_patch.json()
        assert patched_invoice["remaining_amount"] == 0
        assert patched_invoice["status"] == "✅入金済"

        payment_create = client.post(
            "/api/v1/payments",
            json={
                "project_id": "P-003",
                "vendor_name": "テスト業者",
                "ordered_amount": 120000,
                "paid_amount": 20000,
                "status": "❌未支払",
            },
        )
        assert payment_create.status_code == 201
        payment_body = payment_create.json()
        assert payment_body["payment_id"].startswith("PAY-")
        assert payment_body["remaining_amount"] == 100000

        payment_list = client.get("/api/v1/payments", params={"project_id": "P-003"})
        assert payment_list.status_code == 200
        payments = payment_list.json()
        assert any(x["payment_id"] == payment_body["payment_id"] for x in payments)

        payment_patch = client.patch(
            f"/api/v1/payments/{payment_body['payment_id']}",
            json={"paid_amount": 120000},
        )
        assert payment_patch.status_code == 200
        patched_payment = payment_patch.json()
        assert patched_payment["remaining_amount"] == 0
        assert patched_payment["status"] == "✅支払済"


def test_dashboard_summary() -> None:
    with TestClient(app) as client:
        resp = client.get("/api/v1/dashboard/summary")
        assert resp.status_code == 200
        body = resp.json()
        assert body["project_total"] >= 1
        assert isinstance(body["project_status_counts"], dict)
        assert body["invoice_total_amount"] >= 0
        assert body["payment_total_amount"] >= 0
