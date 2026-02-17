"""Excel workbook synchronization service for MVP."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Customer, Invoice, Payment, Project, WorkItemMaster
from .sanitize import sanitize_sheet_name


@dataclass
class SyncResult:
    workbook_path: str
    customers_upserted: int
    projects_upserted: int
    invoices_upserted: int
    payments_upserted: int
    work_items_upserted: int


def _to_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _to_float(value, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(value[:10], fmt).date()
            except ValueError:
                continue
    return None


def _get_or_create_placeholder_customer(db: Session) -> Customer:
    customer = db.execute(select(Customer).where(Customer.customer_id == "C-000")).scalar_one_or_none()
    if customer:
        return customer
    customer = Customer(
        customer_id="C-000",
        customer_name="未設定顧客",
        contact_name=None,
        status="一時",
    )
    db.add(customer)
    db.flush()
    return customer


def sync_from_workbook(db: Session, workbook_path: str) -> SyncResult:
    source = Path(workbook_path).expanduser().resolve()
    if not source.exists():
        raise FileNotFoundError(f"Workbook not found: {source}")

    wb = load_workbook(source, data_only=True, keep_vba=True)

    customers_upserted = 0
    projects_upserted = 0
    invoices_upserted = 0
    payments_upserted = 0
    work_items_upserted = 0

    # 1) 顧客マスタ
    if "顧客マスタ" in wb.sheetnames:
        ws = wb["顧客マスタ"]
        for row in range(5, ws.max_row + 1):
            customer_id = _to_str(ws.cell(row, 1).value)
            customer_name = _to_str(ws.cell(row, 3).value)
            if not customer_id or not customer_name:
                continue
            existing = db.execute(
                select(Customer).where(Customer.customer_id == customer_id)
            ).scalar_one_or_none()
            if existing is None:
                existing = Customer(
                    customer_id=customer_id,
                    customer_name=customer_name,
                )
                db.add(existing)
            existing.customer_name = customer_name
            existing.contact_name = _to_str(ws.cell(row, 5).value)
            existing.status = _to_str(ws.cell(row, 18).value) or "アクティブ"
            customers_upserted += 1

    db.flush()
    placeholder = _get_or_create_placeholder_customer(db)

    # 2) 案件管理
    if "案件管理" in wb.sheetnames:
        ws = wb["案件管理"]
        for row in range(5, ws.max_row + 1):
            project_id = _to_str(ws.cell(row, 1).value)
            if not project_id or not project_id.startswith("P-"):
                continue

            customer_id = _to_str(ws.cell(row, 2).value) or placeholder.customer_id
            customer_name = _to_str(ws.cell(row, 3).value) or placeholder.customer_name
            project_name = _to_str(ws.cell(row, 4).value) or _to_str(ws.cell(row, 32).value) or "案件未設定"
            project_sheet_name = sanitize_sheet_name(f"{project_id}_{project_name}", f"{project_id}_案件")

            customer = db.execute(
                select(Customer).where(Customer.customer_id == customer_id)
            ).scalar_one_or_none()
            if customer is None:
                customer = Customer(
                    customer_id=customer_id,
                    customer_name=customer_name,
                    status="アクティブ",
                )
                db.add(customer)

            existing = db.execute(
                select(Project).where(Project.project_id == project_id)
            ).scalar_one_or_none()
            if existing is None:
                existing = Project(
                    project_id=project_id,
                    project_sheet_name=project_sheet_name,
                    customer_id=customer_id,
                    customer_name=customer_name,
                    project_name=project_name,
                    project_status="①リード",
                )
                db.add(existing)

            existing.project_sheet_name = project_sheet_name
            existing.customer_id = customer_id
            existing.customer_name = customer_name
            existing.project_name = project_name
            existing.site_address = _to_str(ws.cell(row, 31).value)
            existing.owner_name = _to_str(ws.cell(row, 13).value) or "吉野博"
            existing.target_margin_rate = _to_float(ws.cell(row, 8).value, 0.25) or 0.25
            existing.project_status = _to_str(ws.cell(row, 9).value) or "①リード"
            existing.created_at = _to_date(ws.cell(row, 19).value) or date.today()
            projects_upserted += 1

    db.flush()

    # 3) 請求管理
    if "請求管理" in wb.sheetnames:
        ws = wb["請求管理"]
        for row in range(5, ws.max_row + 1):
            project_id = _to_str(ws.cell(row, 2).value)
            if not project_id:
                continue

            invoice_id = _to_str(ws.cell(row, 1).value)
            if not invoice_id or invoice_id.startswith("="):
                invoice_id = f"INV-{row - 4:03d}"

            project = db.execute(
                select(Project).where(Project.project_id == project_id)
            ).scalar_one_or_none()
            if project is None:
                project = Project(
                    project_id=project_id,
                    project_sheet_name=sanitize_sheet_name(project_id, f"{project_id}_案件"),
                    customer_id=placeholder.customer_id,
                    customer_name=placeholder.customer_name,
                    project_name=_to_str(ws.cell(row, 3).value) or project_id,
                    project_status="①リード",
                    created_at=date.today(),
                )
                db.add(project)
                db.flush()

            existing = db.execute(
                select(Invoice).where(Invoice.invoice_id == invoice_id)
            ).scalar_one_or_none()
            if existing is None:
                existing = Invoice(
                    invoice_id=invoice_id,
                    project_id=project_id,
                    invoice_amount=0.0,
                )
                db.add(existing)

            existing.project_id = project_id
            existing.invoice_amount = _to_float(ws.cell(row, 7).value, 0.0)
            existing.billed_at = _to_date(ws.cell(row, 5).value) or existing.billed_at or date.today()
            existing.note = _to_str(ws.cell(row, 12).value)
            invoices_upserted += 1

    # 4) 工事項目DB
    if "工事項目DB" in wb.sheetnames:
        ws = wb["工事項目DB"]
        for row in range(5, ws.max_row + 1):
            category = _to_str(ws.cell(row, 2).value)
            item_name = _to_str(ws.cell(row, 3).value)
            if not category or not item_name:
                continue

            source_item_id = ws.cell(row, 1).value
            source_item_id_int = int(source_item_id) if isinstance(source_item_id, (int, float)) else None

            existing = None
            if source_item_id_int is not None:
                existing = db.execute(
                    select(WorkItemMaster).where(WorkItemMaster.source_item_id == source_item_id_int)
                ).scalar_one_or_none()
            if existing is None:
                existing = db.execute(
                    select(WorkItemMaster).where(
                        WorkItemMaster.category == category,
                        WorkItemMaster.item_name == item_name,
                    )
                ).scalar_one_or_none()

            if existing is None:
                existing = WorkItemMaster(
                    source_item_id=source_item_id_int,
                    category=category,
                    item_name=item_name,
                    standard_unit_price=0.0,
                )
                db.add(existing)

            existing.source_item_id = source_item_id_int
            existing.category = category
            existing.item_name = item_name
            existing.specification = _to_str(ws.cell(row, 4).value)
            existing.unit = _to_str(ws.cell(row, 5).value)
            existing.standard_unit_price = _to_float(ws.cell(row, 6).value, 0.0)
            existing.default_vendor_name = _to_str(ws.cell(row, 7).value)
            margin = ws.cell(row, 9).value
            existing.margin_rate = _to_float(margin, 0.0) if margin is not None else None
            work_items_upserted += 1

    # 5) 支払管理
    if "支払管理" in wb.sheetnames:
        ws = wb["支払管理"]
        for row in range(5, ws.max_row + 1):
            project_id = _to_str(ws.cell(row, 2).value)
            if not project_id:
                continue

            payment_id = _to_str(ws.cell(row, 1).value)
            if not payment_id or payment_id.startswith("="):
                payment_id = f"PAY-{row - 4:03d}"

            project = db.execute(
                select(Project).where(Project.project_id == project_id)
            ).scalar_one_or_none()
            if project is None:
                project = Project(
                    project_id=project_id,
                    project_sheet_name=sanitize_sheet_name(project_id, f"{project_id}_案件"),
                    customer_id=placeholder.customer_id,
                    customer_name=placeholder.customer_name,
                    project_name=_to_str(ws.cell(row, 5).value) or project_id,
                    project_status="①リード",
                    created_at=date.today(),
                )
                db.add(project)
                db.flush()

            existing = db.execute(
                select(Payment).where(Payment.payment_id == payment_id)
            ).scalar_one_or_none()
            if existing is None:
                existing = Payment(
                    payment_id=payment_id,
                    project_id=project_id,
                )
                db.add(existing)

            ordered_amount = _to_float(ws.cell(row, 7).value, 0.0)
            paid_amount = _to_float(ws.cell(row, 9).value, 0.0)
            remaining_amount = _to_float(ws.cell(row, 10).value, ordered_amount - paid_amount)
            if remaining_amount < 0:
                remaining_amount = 0.0

            existing.project_id = project_id
            existing.vendor_id = _to_str(ws.cell(row, 3).value)
            existing.vendor_name = _to_str(ws.cell(row, 4).value)
            existing.work_description = _to_str(ws.cell(row, 5).value)
            existing.ordered_amount = ordered_amount
            existing.paid_amount = paid_amount
            existing.remaining_amount = remaining_amount
            existing.status = _to_str(ws.cell(row, 11).value)
            existing.note = _to_str(ws.cell(row, 13).value)
            existing.paid_at = _to_date(ws.cell(row, 8).value)
            payments_upserted += 1

    db.commit()

    return SyncResult(
        workbook_path=str(source),
        customers_upserted=customers_upserted,
        projects_upserted=projects_upserted,
        invoices_upserted=invoices_upserted,
        payments_upserted=payments_upserted,
        work_items_upserted=work_items_upserted,
    )
